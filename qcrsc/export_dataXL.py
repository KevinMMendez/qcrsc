import pandas as pd
import openpyxl


def export_dataXL(file, DataSheet, PeakSheet, data_sheet='DataTableX', peak_sheet='PeakTableX'):

    # Check if file already exists or create it
    try:
        wb = openpyxl.load_workbook(file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.save(file)

    # Excel sheets
    wb_sheets = wb.sheetnames

    # Delete data_sheet is it's in excel already
    if data_sheet in wb_sheets:
        wb.remove(wb.get_sheet_by_name(data_sheet))

    # Delete peak_sheet is it's in excel already
    if peak_sheet in wb_sheets:
        wb.remove(wb.get_sheet_by_name(peak_sheet))

    # Save excel file
    wb.save(file)

    # Save new sheets
    with pd.ExcelWriter(file, engine='openpyxl', mode='a') as writer:
        DataSheet.to_excel(writer, sheet_name=data_sheet, index=False)
        PeakSheet.to_excel(writer, sheet_name=peak_sheet, index=False)

    print('Done.')
